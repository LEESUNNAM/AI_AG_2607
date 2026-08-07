"""
General-purpose .xlsx workbook builder using openpyxl.

Styling defaults match this project's other document scripts
(scripts/create_docx.py): Malgun Gothic font, a customizable main
color applied to titles and table headers.

Usage as a library:
    from create_xlsx import SheetBuilder

    wb = SheetBuilder(main_color="2E86AB")
    wb.add_title("Sales Report")
    wb.write_table(
        headers=["Name", "Score"],
        rows=[["Alice", 90], ["Bob", 85]],
    )
    wb.save("output/report.xlsx")

Usage from the command line (creates a demo workbook):
    python create_xlsx.py output/demo.xlsx --color 2E86AB --title "Sample Workbook"
"""

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_FONT = "맑은 고딕"
DEFAULT_TITLE_SIZE = 18
DEFAULT_HEADER_SIZE = 11
DEFAULT_BODY_SIZE = 10
DEFAULT_MAIN_COLOR = "1F4E79"


def _normalize_color(color: str) -> str:
    return color.lstrip("#").upper()


class SheetBuilder:
    def __init__(
        self,
        title: str | None = None,
        main_color: str = DEFAULT_MAIN_COLOR,
        font_name: str = DEFAULT_FONT,
        title_size: float = DEFAULT_TITLE_SIZE,
        body_size: float = DEFAULT_BODY_SIZE,
    ):
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.main_color = _normalize_color(main_color)
        self.font_name = font_name
        self.title_size = title_size
        self.body_size = body_size
        self._next_row = 1
        if title:
            self.sheet.title = title[:31]  # Excel sheet name limit
            self.add_title(title)

    def add_sheet(self, name: str):
        self.sheet = self.workbook.create_sheet(title=name[:31])
        self._next_row = 1
        return self

    def add_title(self, text: str, span_cols: int = 6):
        row = self._next_row
        self.sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
        cell = self.sheet.cell(row=row, column=1, value=text)
        cell.font = Font(name=self.font_name, size=self.title_size, bold=True, color=self.main_color)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        self.sheet.row_dimensions[row].height = self.title_size * 1.6
        self._next_row = row + 2
        return self

    def write_table(
        self,
        headers: list[str],
        rows: list[list],
        header_color: str | None = None,
        start_col: int = 1,
    ):
        header_fill = PatternFill(
            start_color=_normalize_color(header_color) if header_color else self.main_color,
            end_color=_normalize_color(header_color) if header_color else self.main_color,
            fill_type="solid",
        )
        header_row = self._next_row
        for i, header in enumerate(headers):
            cell = self.sheet.cell(row=header_row, column=start_col + i, value=header)
            cell.font = Font(name=self.font_name, size=self.body_size, bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r, row_values in enumerate(rows, start=1):
            for c, value in enumerate(row_values):
                cell = self.sheet.cell(row=header_row + r, column=start_col + c, value=value)
                cell.font = Font(name=self.font_name, size=self.body_size)
                cell.alignment = Alignment(horizontal="left", vertical="center")

        self._autofit_columns(headers, rows, start_col)
        self._next_row = header_row + len(rows) + 2
        return self

    def _autofit_columns(self, headers: list[str], rows: list[list], start_col: int):
        for i, header in enumerate(headers):
            values = [str(header)] + [str(row[i]) for row in rows if i < len(row)]
            width = max(len(v) for v in values) + 4
            self.sheet.column_dimensions[get_column_letter(start_col + i)].width = width

    def save(self, path: str):
        out_path = Path(path)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(out_path)
        return out_path


def _demo(path: str, title: str, main_color: str):
    wb = SheetBuilder(title=title, main_color=main_color)
    wb.write_table(
        headers=["Item", "Category", "Value"],
        rows=[
            ["Alpha", "A", 100],
            ["Beta", "B", 85],
            ["Gamma", "A", 42],
        ],
    )
    saved_path = wb.save(path)
    print(f"Saved: {saved_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate a demo .xlsx file.")
    parser.add_argument("output", nargs="?", default="output/demo.xlsx")
    parser.add_argument("--title", default="Sample Workbook")
    parser.add_argument("--color", default=DEFAULT_MAIN_COLOR, help="Main color as a hex string, e.g. 2E86AB")
    args = parser.parse_args()
    _demo(args.output, args.title, args.color)
