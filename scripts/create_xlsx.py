"""
General-purpose .xlsx workbook builder using openpyxl.

Styling defaults match this project's other document scripts
(scripts/create_docx.py): Malgun Gothic font, a customizable main
color applied to titles and table headers.

On top of the basic builder, every workbook produced by this script
(or passed through ``format_workbook``) automatically gets:
  - a leading "ID" column with sequential row numbers
  - bold titles/headers, italic quotes
  - a normalized 10pt body font (existing font attributes preserved)
  - an auto-generated chart when growth-rate / comparison data is found

Usage as a library:
    from create_xlsx import SheetBuilder

    wb = SheetBuilder(main_color="2E86AB")
    wb.add_title("Sales Report")
    wb.write_table(
        headers=["Name", "Score"],
        rows=[["Alice", 90], ["Bob", 85]],
    )
    wb.save("output/report.xlsx")

Usage from the command line (creates a demo workbook):
    python create_xlsx.py output/demo.xlsx --color 2E86AB --title "Sample Workbook"

Usage to reformat an existing .xlsx file (adds ID column, styling, charts):
    python create_xlsx.py --input existing.xlsx --output existing_formatted.xlsx
"""

import argparse
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

DEFAULT_FONT = "맑은 고딕"
DEFAULT_TITLE_SIZE = 18
DEFAULT_HEADER_SIZE = 11
DEFAULT_BODY_SIZE = 10
DEFAULT_MAIN_COLOR = "1F4E79"

# --- keyword tables used by the auto-formatting pass -----------------------

TITLE_KEYWORD_COLUMNS = {"제목", "항목명", "구분", "title", "subject"}
QUOTE_KEYWORD_COLUMNS = {"인용구", "인용문", "quote", "quotation"}
RATE_KEYWORDS = [
    "성장률", "증가율", "증감률", "변화율", "달성률", "점유율", "비율",
    "전년 대비", "전월 대비", "growth rate", "change rate", "rate",
]
SHARE_KEYWORDS = ["점유율", "구성비", "비율", "share"]
TIME_KEYWORDS = ["연도", "년도", "year", "월", "month", "분기", "quarter"]
TOTAL_ROW_KEYWORDS = ["합계", "총계", "소계", "total", "sum"]


def _normalize_color(color: str) -> str:
    return color.lstrip("#").upper()


# --- small shared helpers ---------------------------------------------------

def _set_font(cell, *, bold=None, italic=None, size=None, name=None):
    """Copy a cell's existing Font and change only the given attributes,
    so color/underline/etc. set elsewhere are never clobbered."""
    if isinstance(cell, MergedCell):
        return
    current = cell.font
    new_font = copy(current)
    if bold is not None:
        new_font.bold = bold
    if italic is not None:
        new_font.italic = italic
    if size is not None:
        new_font.sz = size
    if name is not None and not current.name:
        new_font.name = name
    cell.font = new_font


def _find_header_row(ws):
    """Return the 1-indexed row number of the first row with 2+ non-empty
    cells (a merged single-cell title row doesn't count), or None."""
    for row in ws.iter_rows():
        values = [c.value for c in row if c.value not in (None, "")]
        if len(values) >= 2:
            return row[0].row
    return None


def _columns_matching(ws, header_row, keyword_set):
    cols = []
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        if isinstance(value, str) and value.strip().lower() in keyword_set:
            cols.append(col)
    return cols


def is_quote(value) -> bool:
    """Decide whether a cell's string value looks like a quotation."""
    if not isinstance(value, str):
        return False
    s = value.strip()
    if not s:
        return False
    if s.startswith(">"):
        return True
    quote_pairs = [('"', '"'), ("'", "'"), ("“", "”"), ("‘", "’")]
    for start, end in quote_pairs:
        if len(s) >= 2 and s.startswith(start) and s.endswith(end):
            return True
    return False


# --- formatting passes -------------------------------------------------------

def add_id_column(ws, header_row=None):
    """Insert a leading 'ID' column with sequential numbers for each
    non-blank data row, shifting existing data one column to the right.
    Safe to call on sheets that already have an ID column (no-op)."""
    try:
        if ws.max_row < 1 or ws.max_column < 1:
            return
        header_row = header_row or _find_header_row(ws)
        if header_row is None:
            return
        header_cell = ws.cell(row=header_row, column=1)
        if isinstance(header_cell.value, str) and header_cell.value.strip().upper() == "ID":
            return  # already has an ID column

        max_col = ws.max_column
        max_row = ws.max_row

        merged_ranges = [r for r in ws.merged_cells.ranges if r.min_row >= header_row]
        for merged in merged_ranges:
            ws.unmerge_cells(str(merged))

        # shift column widths right by one
        old_widths = {
            col: ws.column_dimensions[get_column_letter(col)].width
            for col in range(1, max_col + 1)
            if get_column_letter(col) in ws.column_dimensions
        }
        for col in range(max_col, 0, -1):
            if col in old_widths:
                ws.column_dimensions[get_column_letter(col + 1)].width = old_widths[col]

        # shift cell contents/styles right by one, within the table's rows only
        for row in range(max_row, header_row - 1, -1):
            for col in range(max_col, 0, -1):
                src = ws.cell(row=row, column=col)
                dst = ws.cell(row=row, column=col + 1)
                dst.value = src.value
                dst.font = copy(src.font)
                dst.fill = copy(src.fill)
                dst.alignment = copy(src.alignment)
                dst.number_format = src.number_format
                dst.border = copy(src.border)
                src.value = None

        for merged in merged_ranges:
            ws.merge_cells(
                start_row=merged.min_row, start_column=merged.min_col + 1,
                end_row=merged.max_row, end_column=merged.max_col + 1,
            )

        neighbor_header = ws.cell(row=header_row, column=2)
        id_header = ws.cell(row=header_row, column=1, value="ID")
        id_header.font = copy(neighbor_header.font)
        id_header.fill = copy(neighbor_header.fill)
        id_header.alignment = Alignment(horizontal="center", vertical="center")

        idx = 1
        for row in range(header_row + 1, max_row + 1):
            row_values = [ws.cell(row=row, column=c).value for c in range(2, max_col + 2)]
            if all(v in (None, "") for v in row_values):
                continue  # don't number blank rows
            cell = ws.cell(row=row, column=1, value=idx)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            idx += 1

        ws.column_dimensions["A"].width = max(6, len(str(idx - 1)) + 4)
    except Exception as e:
        print(f"[경고] ID 열 추가 실패 (시트 '{ws.title}'): {e}")


def apply_header_style(ws, header_row=None, font_name=DEFAULT_FONT):
    """Bold the header row, any merged report-title cells above it, and any
    columns whose header matches a title keyword (제목/항목명/구분/...)."""
    try:
        header_row = header_row or _find_header_row(ws)
        if header_row is None:
            return
        for cell in ws[header_row]:
            if cell.value in (None, ""):
                continue
            _set_font(cell, bold=True, name=font_name)

        for merged in ws.merged_cells.ranges:
            if merged.max_row < header_row and (merged.max_col - merged.min_col) >= 1:
                top_left = ws.cell(row=merged.min_row, column=merged.min_col)
                _set_font(top_left, bold=True, name=font_name)

        for col in _columns_matching(ws, header_row, TITLE_KEYWORD_COLUMNS):
            for row in range(header_row + 1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if not isinstance(cell, MergedCell) and cell.value not in (None, ""):
                    _set_font(cell, bold=True, name=font_name)
    except Exception as e:
        print(f"[경고] 헤더/제목 서식 적용 실패 (시트 '{ws.title}'): {e}")


def apply_quote_style(ws, header_row=None):
    """Italicize cells recognized as quotations, either by column label
    (인용구/인용문/Quote/Quotation) or by quote-like punctuation."""
    try:
        header_row = header_row or _find_header_row(ws)
        if header_row is None:
            return
        quote_cols = set(_columns_matching(ws, header_row, QUOTE_KEYWORD_COLUMNS))
        for row in ws.iter_rows(min_row=header_row + 1):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if cell.column in quote_cols or is_quote(cell.value):
                    _set_font(cell, italic=True)
    except Exception as e:
        print(f"[경고] 인용구 서식 적용 실패 (시트 '{ws.title}'): {e}")


def apply_body_font(ws, header_row=None, font_name=DEFAULT_FONT, body_size=DEFAULT_BODY_SIZE):
    """Normalize body data (everything except header row and title-keyword
    columns) to body_size, preserving bold/italic/color/etc."""
    try:
        header_row = header_row or _find_header_row(ws)
        if header_row is None:
            return
        title_cols = set(_columns_matching(ws, header_row, TITLE_KEYWORD_COLUMNS))
        for row in ws.iter_rows(min_row=header_row + 1):
            for cell in row:
                if isinstance(cell, MergedCell) or cell.value in (None, ""):
                    continue
                if cell.column in title_cols:
                    continue
                _set_font(cell, size=body_size, name=font_name)
    except Exception as e:
        print(f"[경고] 본문 글꼴 적용 실패 (시트 '{ws.title}'): {e}")


def detect_numeric_columns(ws, header_row=None):
    """Inspect header labels and data to find a category column plus any
    numeric / growth-rate columns usable for a chart."""
    result = {"category_col": None, "numeric_cols": [], "rate_cols": [], "time_like": False}
    try:
        header_row = header_row or _find_header_row(ws)
        if header_row is None:
            return result
        max_col = ws.max_column
        headers = {
            col: (str(ws.cell(row=header_row, column=col).value).strip()
                  if ws.cell(row=header_row, column=col).value is not None else "")
            for col in range(1, max_col + 1)
        }
        data_rows = [
            r for r in range(header_row + 1, ws.max_row + 1)
            if any(ws.cell(row=r, column=c).value not in (None, "") for c in range(1, max_col + 1))
        ]
        if not data_rows:
            return result

        start_col = 2 if headers.get(1, "").upper() == "ID" else 1
        for col in range(start_col, max_col + 1):
            header_text = headers[col]
            values = []
            numeric_ok = True
            for r in data_rows:
                v = ws.cell(row=r, column=col).value
                if v in (None, ""):
                    continue
                if isinstance(v, str) and v.startswith("="):
                    numeric_ok = False
                    break
                try:
                    values.append(float(str(v).replace("%", "").replace(",", "")))
                except (TypeError, ValueError):
                    numeric_ok = False
                    break
            if numeric_ok and values:
                result["numeric_cols"].append(col)
                if any(kw.lower() in header_text.lower() for kw in RATE_KEYWORDS):
                    result["rate_cols"].append(col)
            elif result["category_col"] is None and header_text:
                result["category_col"] = col

        if result["category_col"] is None and start_col <= max_col:
            result["category_col"] = start_col

        cat_header = headers.get(result["category_col"], "") if result["category_col"] else ""
        result["time_like"] = any(kw.lower() in cat_header.lower() for kw in TIME_KEYWORDS)
    except Exception as e:
        print(f"[경고] 숫자 열 탐지 실패 (시트 '{ws.title}'): {e}")
    return result


def _apply_percentage_format(ws, header_row, data_max_row, rate_cols):
    try:
        for col in rate_cols:
            values = [
                ws.cell(row=r, column=col).value
                for r in range(header_row + 1, data_max_row + 1)
                if isinstance(ws.cell(row=r, column=col).value, (int, float))
            ]
            if not values:
                continue
            fmt = "0.0%" if all(-1 <= v <= 1 for v in values) else '0.0"%"'
            for r in range(header_row + 1, data_max_row + 1):
                cell = ws.cell(row=r, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = fmt
    except Exception as e:
        print(f"[경고] 퍼센트 서식 적용 실패 (시트 '{ws.title}'): {e}")


def create_chart_if_needed(ws, header_row=None):
    """Auto-generate a Line/Bar/Pie chart when the sheet has growth-rate or
    comparable numeric data. Returns the created chart, or None."""
    try:
        header_row = header_row or _find_header_row(ws)
        if header_row is None:
            return None

        info = detect_numeric_columns(ws, header_row)
        numeric_cols = info["numeric_cols"]
        category_col = info["category_col"]
        if not numeric_cols or category_col is None:
            return None

        data_max_row = ws.max_row
        relevant_cols = [category_col] + numeric_cols
        while data_max_row > header_row and all(
            ws.cell(row=data_max_row, column=c).value in (None, "") for c in relevant_cols
        ):
            data_max_row -= 1
        while data_max_row > header_row:
            cat_val = ws.cell(row=data_max_row, column=category_col).value
            if isinstance(cat_val, str) and any(k in cat_val.lower() for k in TOTAL_ROW_KEYWORDS):
                data_max_row -= 1
            else:
                break
        if data_max_row <= header_row:
            return None

        _apply_percentage_format(ws, header_row, data_max_row, info["rate_cols"])

        category_header = str(ws.cell(row=header_row, column=category_col).value or "구분")

        if info["rate_cols"] or info["time_like"]:
            chart = LineChart()
            value_cols = info["rate_cols"] or numeric_cols
            chart.title = f"{category_header} 변화 추이"
        else:
            headers_text = [str(ws.cell(row=header_row, column=c).value or "") for c in numeric_cols]
            is_share = any(any(k in h.lower() for k in SHARE_KEYWORDS) for h in headers_text)
            n_categories = data_max_row - header_row
            if is_share and len(numeric_cols) == 1 and n_categories <= 7:
                chart = PieChart()
                value_cols = numeric_cols
                chart.title = f"{headers_text[0]} 구성비"
            elif len(numeric_cols) >= 1:
                chart = BarChart()
                value_cols = numeric_cols
                chart.title = "항목별 수치 비교"
            else:
                return None

        chart.width = 14
        chart.height = 8
        if not isinstance(chart, PieChart):
            chart.x_axis.title = category_header
            chart.y_axis.title = "값"

        data_ref = Reference(ws, min_col=min(value_cols), max_col=max(value_cols),
                              min_row=header_row, max_row=data_max_row)
        cats_ref = Reference(ws, min_col=category_col, max_col=category_col,
                              min_row=header_row + 1, max_row=data_max_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        existing_charts = getattr(ws, "_charts", [])
        anchor_col = ws.max_column + 2
        anchor_row = 2 + len(existing_charts) * 17
        ws.add_chart(chart, f"{get_column_letter(anchor_col)}{anchor_row}")
        return chart
    except Exception as e:
        print(f"[경고] 차트 생성 실패 (시트 '{ws.title}'): {e}")
        return None


def format_workbook(workbook, font_name: str = DEFAULT_FONT):
    """Apply the full formatting pass (ID column, header/quote/body styling,
    auto chart) to every sheet in a workbook. Never raises."""
    try:
        sheets = workbook.worksheets
    except Exception as e:
        print(f"[경고] 워크북에서 시트를 읽을 수 없습니다: {e}")
        return
    if not sheets:
        print("[경고] 워크북에 시트가 없습니다.")
        return

    for ws in sheets:
        try:
            if ws.max_row < 1 or ws.max_column < 1 or ws.cell(row=1, column=1).value is None and ws.max_row == 1 and ws.max_column == 1:
                continue  # empty sheet
            header_row = _find_header_row(ws)
            if header_row is None:
                continue  # can't locate a header row; leave sheet untouched
            add_id_column(ws, header_row)
            header_row = _find_header_row(ws)
            apply_header_style(ws, header_row, font_name)
            apply_quote_style(ws, header_row)
            apply_body_font(ws, header_row, font_name)
            create_chart_if_needed(ws, header_row)
        except Exception as e:
            print(f"[경고] 시트 '{ws.title}' 서식 적용 중 오류가 발생해 건너뜁니다: {e}")


class SheetBuilder:
    def __init__(
        self,
        title: str | None = None,
        main_color: str = DEFAULT_MAIN_COLOR,
        font_name: str = DEFAULT_FONT,
        title_size: float = DEFAULT_TITLE_SIZE,
        body_size: float = DEFAULT_BODY_SIZE,
    ):
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.main_color = _normalize_color(main_color)
        self.font_name = font_name
        self.title_size = title_size
        self.body_size = body_size
        self._next_row = 1
        if title:
            self.sheet.title = title[:31]  # Excel sheet name limit
            self.add_title(title)

    def add_sheet(self, name: str):
        self.sheet = self.workbook.create_sheet(title=name[:31])
        self._next_row = 1
        return self

    def add_title(self, text: str, span_cols: int = 6):
        row = self._next_row
        self.sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
        cell = self.sheet.cell(row=row, column=1, value=text)
        cell.font = Font(name=self.font_name, size=self.title_size, bold=True, color=self.main_color)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        self.sheet.row_dimensions[row].height = self.title_size * 1.6
        self._next_row = row + 2
        return self

    def write_table(
        self,
        headers: list[str],
        rows: list[list],
        header_color: str | None = None,
        start_col: int = 1,
    ):
        header_fill = PatternFill(
            start_color=_normalize_color(header_color) if header_color else self.main_color,
            end_color=_normalize_color(header_color) if header_color else self.main_color,
            fill_type="solid",
        )
        header_row = self._next_row
        for i, header in enumerate(headers):
            cell = self.sheet.cell(row=header_row, column=start_col + i, value=header)
            cell.font = Font(name=self.font_name, size=self.body_size, bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r, row_values in enumerate(rows, start=1):
            for c, value in enumerate(row_values):
                cell = self.sheet.cell(row=header_row + r, column=start_col + c, value=value)
                cell.font = Font(name=self.font_name, size=self.body_size)
                cell.alignment = Alignment(horizontal="left", vertical="center")

        self._autofit_columns(headers, rows, start_col)
        self._next_row = header_row + len(rows) + 2
        return self

    def _autofit_columns(self, headers: list[str], rows: list[list], start_col: int):
        for i, header in enumerate(headers):
            values = [str(header)] + [str(row[i]) for row in rows if i < len(row)]
            width = max(len(v) for v in values) + 4
            self.sheet.column_dimensions[get_column_letter(start_col + i)].width = width

    def save(self, path: str):
        try:
            format_workbook(self.workbook, self.font_name)
        except Exception as e:
            print(f"[경고] 서식 자동 적용 중 오류가 발생했습니다: {e}")

        out_path = Path(path)
        try:
            out_path.parent.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            print(f"저장 경로를 생성할 수 없습니다: {out_path.parent}\n{e}")
            raise

        try:
            self.workbook.save(out_path)
        except PermissionError:
            print(
                f"저장할 수 없습니다: {out_path}\n"
                "엑셀 파일이 열려 있는지 확인하세요."
            )
            raise
        except Exception as e:
            print(f"저장 중 오류가 발생했습니다: {out_path}\n{e}")
            raise
        return out_path


def load_and_format(input_path: str, output_path: str | None = None):
    """Load an existing .xlsx file, apply the full formatting pass, and
    save it. Handles the common failure modes without crashing."""
    in_path = Path(input_path)
    if not in_path.exists():
        print(f"입력 파일을 찾을 수 없습니다: {in_path}")
        return None

    try:
        workbook = load_workbook(in_path)
    except InvalidFileException as e:
        print(f"엑셀 파일로 인식할 수 없습니다: {in_path}\n{e}")
        return None
    except Exception as e:
        print(f"파일을 여는 중 오류가 발생했습니다: {in_path}\n{e}")
        return None

    format_workbook(workbook)

    out_path = Path(output_path) if output_path else in_path.with_name(f"{in_path.stem}_formatted{in_path.suffix}")
    if out_path.resolve() == in_path.resolve():
        out_path = in_path.with_name(f"{in_path.stem}_formatted{in_path.suffix}")
        print(f"[안내] 입력/출력 경로가 동일해 출력 파일명을 변경했습니다: {out_path.name}")

    try:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(out_path)
    except PermissionError:
        print(f"저장할 수 없습니다: {out_path}\n엑셀 파일이 열려 있는지 확인하세요.")
        return None
    except Exception as e:
        print(f"저장 중 오류가 발생했습니다: {out_path}\n{e}")
        return None

    print(f"Saved: {out_path}")
    return out_path


def _demo(path: str, title: str, main_color: str):
    wb = SheetBuilder(title=title, main_color=main_color)
    wb.write_table(
        headers=["Item", "Category", "Value"],
        rows=[
            ["Alpha", "A", 100],
            ["Beta", "B", 85],
            ["Gamma", "A", 42],
        ],
    )
    saved_path = wb.save(path)
    print(f"Saved: {saved_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate a demo .xlsx file, or reformat an existing one.")
    parser.add_argument("output", nargs="?", default="output/demo.xlsx")
    parser.add_argument("--title", default="Sample Workbook")
    parser.add_argument("--color", default=DEFAULT_MAIN_COLOR, help="Main color as a hex string, e.g. 2E86AB")
    parser.add_argument("--input", default=None, help="Reformat this existing .xlsx instead of building a demo")
    parser.add_argument("--output", dest="format_output", default=None, help="Output path when using --input")
    args = parser.parse_args()

    if args.input:
        load_and_format(args.input, args.format_output)
    else:
        _demo(args.output, args.title, args.color)
